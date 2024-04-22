import logging
import time
import random
import re
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from collections import Counter

# Configure logging
logging.basicConfig(level=logging.DEBUG)  # Set the root logger level to DEBUG

# Create a logger
logger = logging.getLogger(__name__)

# Define log levels to cycle through
log_levels = {
    'INFO': logging.INFO,
    'DEBUG': logging.DEBUG,
    'WARNING': logging.WARNING,
    'ERROR': logging.ERROR
}
# Define keywords or patterns to search for
keywords = ['Database connection failed', 'Data processing completed', 'Application stopped']

# Create a counter for keyword occurrences
keyword_counter = Counter()

# Define lists to store log messages based on log levels
logINFO = []
logDEBUG = []
logWARNING = []
logERROR = []

# Read log messages from logs.txt and parse them
log_messages = []
with open('./logs.txt', 'r') as logfile:
    for line in logfile:
        parts = re.split(r'\s+', line.strip())
        if len(parts) >= 4 and parts[2] in log_levels:
            timestamp, level, message = ' '.join(parts[:2]), parts[2], ' '.join(parts[3:])
            log_messages.append((timestamp, log_levels[level], message))

# Main loop to log messages
try:
    while True:
        # Randomly select a log message from logs.txt
        timestamp, log_level, log_message = random.choice(log_messages)

        # Log the message
        logger.log(log_level, f"{timestamp}  {log_message}")

        # Append to the respective log level list
        if log_level == logging.INFO:
            logINFO.append((timestamp, log_level, log_message))
        elif log_level == logging.DEBUG:
            logDEBUG.append((timestamp, log_level, log_message))
        elif log_level == logging.WARNING:
            logWARNING.append((timestamp, log_level, log_message))
        elif log_level == logging.ERROR:
            logERROR.append((timestamp, log_level, log_message))
        for keyword in keywords:
            if keyword in log_message:
                keyword_counter[keyword] += 1

        # Convert lists to Pandas DataFrames
        df_info = pd.DataFrame(logINFO, columns=['Timestamp', 'Level', 'Message'])
        df_debug = pd.DataFrame(logDEBUG, columns=['Timestamp', 'Level', 'Message'])
        df_warning = pd.DataFrame(logWARNING, columns=['Timestamp', 'Level', 'Message'])
        df_error = pd.DataFrame(logERROR, columns=['Timestamp', 'Level', 'Message'])

        # Add a numbered column
        df_info.insert(0, 'Number', range(1, len(df_info) + 1))
        df_debug.insert(0, 'Number', range(1, len(df_debug) + 1))
        df_warning.insert(0, 'Number', range(1, len(df_warning) + 1))
        df_error.insert(0, 'Number', range(1, len(df_error) + 1))

        # Export DataFrames to Excel
        excel_file_path = 'log_data.xlsx'
        with pd.ExcelWriter(excel_file_path) as writer:
            df_info.to_excel(writer, sheet_name='Info', index=False)
            df_debug.to_excel(writer, sheet_name='Debug', index=False)
            df_warning.to_excel(writer, sheet_name='Warning', index=False)
            df_error.to_excel(writer, sheet_name='Error', index=False)

            # Get the total counts
            total_info = len(df_info)
            total_debug = len(df_debug)
            total_warning = len(df_warning)
            total_error = len(df_error)

            # Create a summary DataFrame
            df_summary = pd.DataFrame({
                'Level': ['INFO', 'DEBUG', 'WARNING', 'ERROR', 'TOTAL'],
                'Count': [total_info, total_debug, total_warning, total_error, total_info + total_debug + total_warning + total_error]
            })

            # Write the summary to Excel
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            df_keywords = pd.DataFrame(keyword_counter.items(), columns=['Keyword', 'Count'])
            df_keywords.to_excel(writer, sheet_name='Keyword Counts', index=False)

        # Modify Excel file formatting
        wb = load_workbook(excel_file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            # Set header row background color
            header_fill = PatternFill(start_color='82acf5', end_color='82acf5', fill_type='solid')
            for cell in ws[1]:
                cell.fill = header_fill

            # Apply alternating row colors white and gray
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if idx % 2 == 0:
                    for cell in row:
                        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                else:
                    for cell in row:
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Save the modified Excel file
        wb.save(excel_file_path)
        
        # Close the Excel workbook
        wb.close()

        # Sleep for a short interval
        time.sleep(1)
except KeyboardInterrupt:
    # Handle keyboard interrupt (Ctrl+C)
    print("\nLogging interrupted. Moving the Excel file.")
    # Move the Excel file to the destination path
    destination_path = r"C:\Users\ASUS\Downloads"
    shutil.move(excel_file_path, destination_path)
    print("Excel file moved successfully. Exiting.")
